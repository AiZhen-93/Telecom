import json
import re
import sys
from pathlib import Path

import openpyxl


WORKBOOK = Path("Aizhen-Telecom測速整理_20260806.xlsx")
OUTPUT = Path("hsr-project-data.js")
STATIONS_NORTH = ["台中", "苗栗", "新竹", "桃園", "板橋", "台北", "南港"]
STATIONS_SOUTH = ["左營", "台南", "嘉義", "雲林", "彰化", "台中"]


def video_id(url):
    if not url:
        return ""
    patterns = [
        r"(?:v=)([A-Za-z0-9_-]{11})",
        r"youtu\.be/([A-Za-z0-9_-]{11})",
        r"youtube\.com/shorts/([A-Za-z0-9_-]{11})",
    ]
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return match.group(1)
    return ""


def clean_date(value):
    match = re.search(r"\d{4}/\d{1,2}/\d{1,2}", str(value or ""))
    if not match:
        return ""
    parts = match.group(0).split("/")
    return f"{int(parts[0]):04d}/{int(parts[1]):02d}/{int(parts[2]):02d}"


def location_group_count(value):
    return len(re.findall(r"\d{3}[^、，,\s]+", str(value or "")))


def route_from_title(title):
    bracket = re.search(r"[［\[]([^］\]]+)[］\]]", title)
    candidates = []
    if bracket:
        candidates.append(bracket.group(1))
    candidates.append(title)
    stations = set(STATIONS_NORTH + STATIONS_SOUTH)
    for text in candidates:
        found = [station for station in stations if station in text]
        if len(found) >= 2:
            ordered = sorted(found, key=lambda name: text.index(name))
            return ordered[0], ordered[-1]
        if len(found) == 1:
            return found[0], found[0]
    return "", ""


def segment_for_route(start, end):
    if start in STATIONS_NORTH and end in STATIONS_NORTH:
        return "north"
    if start in STATIONS_SOUTH and end in STATIONS_SOUTH:
        return "south"
    return ""


def operator_key(value):
    text = str(value or "")
    local_operators = sum(1 for name in ["中華", "遠傳", "台灣大", "台哥", "台灣之星", "台星", "亞太"] if name in text)
    if local_operators >= 2:
        return "multi"
    if "中華" in text:
        return "cht"
    if "遠傳" in text:
        return "fet"
    if "台灣大" in text or "台哥" in text:
        return "twm"
    if "台灣之星" in text or "台星" in text:
        return "tstar"
    if "亞太" in text:
        return "apt"
    return "multi" if any(mark in text for mark in ["、", "/", "+", "與", "和"]) else "unknown"


wb = openpyxl.load_workbook(WORKBOOK, read_only=False, data_only=True)
ws = wb["測試資料"]
headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
idx = {header: col + 1 for col, header in enumerate(headers)}

items = []
for row in range(2, ws.max_row + 1):
    title = str(ws.cell(row, idx["測試影片(點選可觀看)"]).value or "")
    content = str(ws.cell(row, idx["測試內容"]).value or "")
    if "高鐵" not in title or re.search(r"Speed\s*Test", content, re.IGNORECASE):
        continue

    location_count = location_group_count(ws.cell(row, idx["所處行政區"]).value)
    if location_count <= 3:
        continue

    title_cell = ws.cell(row, idx["測試影片(點選可觀看)"])
    url = title_cell.hyperlink.target if title_cell.hyperlink else ""
    start, end = route_from_title(title)
    segment = segment_for_route(start, end)
    vid = video_id(url)
    items.append(
        {
            "row": row,
            "date": clean_date(ws.cell(row, idx["測試時間"]).value),
            "operator": str(ws.cell(row, idx["測試電信"]).value or ""),
            "operatorKey": operator_key(ws.cell(row, idx["測試電信"]).value),
            "content": content,
            "locationCount": location_count,
            "project": str(ws.cell(row, idx["測試專案"]).value or ""),
            "playlist": str(ws.cell(row, idx["所屬播放清單"]).value or ""),
            "title": title,
            "url": url,
            "videoId": vid,
            "thumbnail": f"pic/HSR_{vid}.jpg" if vid else "",
            "start": start,
            "end": end,
            "segment": segment,
        }
    )

items = [item for item in items if item["segment"] and item["videoId"]]
items.sort(key=lambda item: item["date"], reverse=True)

if "--js" in sys.argv:
    payload = {
        "updated": "2026/08/07",
        "stations": {
            "north": STATIONS_NORTH,
            "south": STATIONS_SOUTH,
        },
        "items": items,
    }
    OUTPUT.write_text(
        "window.hsrProjectData = "
        + json.dumps(payload, ensure_ascii=False, indent=4)
        + ";\n",
        encoding="utf-8",
    )
else:
    print(json.dumps(items, ensure_ascii=False, indent=2))
